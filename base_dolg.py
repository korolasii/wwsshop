import openpyxl
import json


def base_dolg_reload(way='Main_new.xlsm'):
    data={}
    book = openpyxl.load_workbook(str(way), data_only=True)  # Открытие файла
    baze_dolg = book.worksheets[book.sheetnames.index('Долги')]
    for i in range(2,baze_dolg.max_row):
        data[baze_dolg[i][0].value.strip()]=round(baze_dolg[i][1].value, 2)
    open_file_json_w(data)
    return data

def open_file_json_w(data):
    with open('base_dolg.json', 'w') as file:
       json.dump(data, file, indent=4)
    return data

if __name__ == '__main__':
    print(base_dolg_reload())