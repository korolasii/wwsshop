import openpyxl
from datetime import datetime
import base_json_test
import base_dolg
import PySimpleGUI as sg
import time



def header_table(way='Main_new.xlsm'):
    book = openpyxl.open(str(way), read_only=True)  # Открытие файла
    sheet = book.worksheets[book.sheetnames.index('получение')]  # Запуск листа с названием - получение
    subject = sheet[3][1].value
    second_start(sheet)
    message_for_header_table(sheet, table_header=20)
    return subject, message,buyers_diferent_test(sheet, table_header), mylist_def(), header_tabel_second_start(sheet), body_table(
        sheet), dolg(base_dolg.base_dolg_reload(way))


def second_start(sheet):
    global table_header
    for i in range(1, sheet.max_row):
        if sheet[i][0].value == 'дата ':
            table_header = i
            break
    return table_header

#Создание шабки таблицы ввиде словаря
def header_tabel_second_start(sheet):
    sg.one_line_progress_meter('Рассылка', mylist[1], len(gmail_list.keys()) * 2 + 3, 'Собераем даные о письме')
    time.sleep(1)
    global message_header_tabel_dickt
    message_header_tabel_dickt={}
    key = ''
    for i in list(*sheet.iter_rows(min_row=second_start(sheet), max_row=second_start(sheet), min_col=0, max_col=8, values_only=True)):
        if i != None:
            key += 'b'
            message_header_tabel_dickt[key] = i
    return message_header_tabel_dickt

def body_table(sheet):
    table_buyer = []
    for i in range(second_start(sheet) + 1, sheet.max_row + 1):
        a = []
        for j in list(*sheet.iter_rows(min_row=i, max_row=i, min_col=0, max_col=8, values_only=True)):
            if isinstance(j, datetime):
                a.append(j.strftime('%d/%m/%y'))
            elif j==None:
                a.append('')
            else:
                a.append(j)
        table_buyer.append(a)
    return table_buyer

def buyers_diferent_test(sheet, table_header):
    global gmail_list
    gmail_baze = base_json_test.base_buer()
    gmail_list = {}
    for i in range(table_header + 1, sheet.max_row+1):
        if sheet[i][2].value.strip() not in gmail_list.keys() and sheet[i][2].value.strip() in gmail_baze.keys():
            gmail_list[sheet[i][2].value.strip()] = gmail_baze[sheet[i][2].value.strip()]
    return gmail_list

def message_for_header_table(file_exel, table_header):
    global message
    message = {}
    key = 'a'
    for i in range(4, table_header):
        if file_exel[i][0].value != None:
            key += 'a'
            message[key] = file_exel[i][0].value


def dolg(dolg_dickt):
    sg.one_line_progress_meter('Рассылка', mylist[2], len(gmail_list.keys()) * 2 + 3, 'Собираем даные о долгах')
    time.sleep(1)
    dolg_list={}
    for key, value in dolg_dickt.items():
        dolg_list[key]=value
    return dolg_list

def mylist_def():
    global mylist
    mylist = [1,2,3]
    for i in range(len(gmail_list.keys())*2):
        mylist.append(mylist[len(mylist)-1]+1)
    sg.one_line_progress_meter('Рассылка', mylist[0], len(gmail_list.keys()) * 2 + 3,'Мы создали базу клиентов')
    time.sleep(1)
    return mylist



if __name__ == '__main__':
    print(header_table('Main_new.xlsm'))

    # v=list(sheet.iter_rows(min_row=21, max_row=21, min_col=0, max_col=8, values_only=True)) Берет строки по указаным индексам
